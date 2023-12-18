import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import undetected_chromedriver as uc
from bs4 import BeautifulSoup
import openpyxl
import time

url = []

workbook = openpyxl.load_workbook('DATA EXTRACTING PROJECT 10-25-23.xlsx')

sheet = workbook['Sheet1']
print('Scanning Excel Sheet')
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
    for cell in row:
        url.append(str(cell.value).replace(' ', ''))
print('Scanning Successfull')

def write_data_to_excel(file_path, sheet_name, data_to_insert, row):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    for column, value in enumerate(data_to_insert, start=1):
        sheet.cell(row=row, column=column, value=value)

    workbook.save(file_path)
    workbook.close()
    print('Saved to Excel Sheet')

Path = "C:/Program Files (x86)/chromedriver.exe"
service = Service(Path)
user_agent = 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'

cho = Options()
cho.add_argument("--disable-dev-shm-usage")
cho.add_argument("disable-infobars")
cho.add_argument("--disable-extensions")
cho.add_argument("--no-sandbox")
cho.add_argument(f"user-agent={user_agent}")

driver = uc.Chrome(options=cho, service=service)


def scrap(url):
    driver.get(url)
    time.sleep(2)
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, 'html.parser')
    try:
        table = soup.find('table', {'class': 'table table-striped'})
        return table
    except Exception as e:
        print(f"Error: {e}")
        return None


def extract_attributes(html):
    if html is None:
        return None 

    soup = BeautifulSoup(html, 'html.parser')

    rows = soup.find_all('tr')
    row_data = []
    if len(rows)==0:
        row_data.append('Appeals Not Found for this Property')
        row_data.append('')
        row_data.append('')
        row_data.append('')
        row_data.append('Appeals Not Found for this Property')
        row_data.append('')
        row_data.append('')
        row_data.append('')
        row_data.append('Appeals Not Found for this Property')
        row_data.append('')
        row_data.append('')
        row_data.append('')
    else:
        for row in rows[1:4]:
            columns = row.find_all('td')
            row_data.append(columns[0].text)
            row_data.append(columns[1].text)
            row_data.append(columns[3].text)
            row_data.append(columns[4].text)

    return row_data

count=1
for z, i in enumerate(url[count-1:5000]):
    print(z, i)
    table_content = scrap(i)
    if table_content is not None:
        row_data = extract_attributes(str(table_content))
        print(row_data)
        write_data_to_excel('new.xlsx', 'Sheet1', row_data, count)
        count+=1
driver.quit()
