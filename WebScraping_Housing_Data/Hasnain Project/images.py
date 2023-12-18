import openpyxl
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import undetected_chromedriver as uc
from bs4 import BeautifulSoup
import requests
import os

cho = Options()

Path = "C:\Projects\chromedriver.exe"

service = Service(Path)
cho.add_argument("--disable-dev-shm-usage")
cho.add_argument("disable-infobars")
cho.add_argument("--disable-extensions")
cho.add_argument("--no-sandbox")

driver = uc.Chrome(options=cho, service=service)


workbook = openpyxl.load_workbook('DATA EXTRACTING PROJECT 10-25-23.xlsx')

sheet = workbook['Sheet1']
extracted_data=[]
print('Scanning Excel Sheet')
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
    row_data=[]
    for cell in row:
        row_data.append(str(cell.value).replace(' ', ''))
    extracted_data.append(row_data)
print('Scanning Successfull')



def write_data_to_excel(file_path, sheet_name, data_to_insert, row):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    for column, value in enumerate(data_to_insert, start=1):
        sheet.cell(row=row, column=column, value=value)
    workbook.save(file_path)
    workbook.close()
    print('Saved to Excel Sheet')

count=1
for i, data in enumerate(extracted_data[count-1:110000]):
    print(i,data)
    driver.get(data[1])
    data_to_insert=[]
    found=0
    while(found==0):
        try:
            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            try:
                danger=soup.find('div',{'class':'alert alert-danger spacer'})
                if danger is not None:
                    break
                try:
                    img_link = soup.find('a', {'class': 'imggallery'})['href']
                    img_link = 'https://lrv.nassaucountyny.gov'+img_link
                    notfound=False
                except:
                    notfound=True
                    img_link='No image'
            except:
                pass
            searching_assed_value = soup.find('div', {'id': 'infovaltab'}).find(
                'table').find('tbody').findAll('tr')[3].findAll('td')[1].text.strip()
            schooltax = soup.find('div', {'id': 'infotaxtab'}).find('table').find('tbody').findAll(
                'tr')
            totalschooltax=''
            for i in schooltax:
                finding_schooltax=i.findAll('td')
                for j in finding_schooltax:
                    if j.text.strip()=='Total School Taxes':
                        totalschooltax=finding_schooltax[3].text.replace(
                            '$', '').replace(',', '').strip()
            data_to_insert.append(data[1])
            data_to_insert.append(img_link)
            if notfound==True:
                data_to_insert.append('No image')
            else:
                data_to_insert.append(f"{data[0]}.jpg")
            data_to_insert.append(searching_assed_value)
            data_to_insert.append(totalschooltax)
            print('Assessed Value:', searching_assed_value)
            print('School Tax 2024:', totalschooltax)
            driver.find_element(By.LINK_TEXT, 'General and School Taxes').click()
            driver.find_element(By.ID, 'selectyr').click()
            finding_general_tax = driver.find_element(
                By.ID, 'selectyr').find_elements(By.TAG_NAME, 'option')
            found=1
        except Exception as e:
            continue
    if danger is not None:
        print('No properties info')
        data_to_insert.append(data[1])
        data_to_insert.append('No properties info')
        write_data_to_excel('new_1.xlsx', 'Sheet1', data_to_insert, count)
        count+=1
        continue
    for i in finding_general_tax:
        if i.get_property('value') == '2023':
            i.click()
            found=0
            while(found==0):
                try:
                    page_source = driver.page_source
                    soup = BeautifulSoup(page_source, 'html.parser')
                    gentax = soup.find('div', {'id': 'gentax2023'}).find(
                        'table').find('tbody').findAll('tr')
                    general_tax=''
                    for j in gentax:
                        finding_total = j.findAll('td')
                        for k in finding_total:
                            if k.text.strip() == 'TOTAL':
                                general_tax = finding_total[4].text.strip().replace(
                                    '$', '').replace(',', '')
                    schooltax2023 = soup.find('div', {'id': 'infotaxtab'}).find('table').find('tbody').findAll(
                        'tr')
                    totalschooltax2023=''
                    for i in schooltax2023:
                        finding_schooltax2023 = i.findAll('td')
                        for j in finding_schooltax2023:
                            if j.text.strip() == 'Total School Taxes':
                                totalschooltax2023 = finding_schooltax2023[3].text.replace(
                                    '$', '').replace(',', '').strip()
                    print('School Tax 2023:',totalschooltax2023)
                    print('General Tax 2023:', general_tax)
                    data_to_insert.append(totalschooltax2023)
                    data_to_insert.append(general_tax)
                    found=1
                    break
                except Exception as e:
                    continue
            break    
    try:
        property_description = soup.find('div', {'id': 'procards1'}).findAll('table')
        for i in property_description:
            finding_propertyinfo=i.find('tbody').findAll('tr')
            for i in finding_propertyinfo:
                property_info = i.find('td').text.strip()
                if property_info=='':
                    data_to_insert.append('0')
                else:
                    try:
                        float(property_info)
                        if property_info.find('.')!=-1:
                            data_to_insert.append(float(property_info))
                        else:
                            data_to_insert.append(int(property_info))
                    except:
                        data_to_insert.append(property_info)
    except:
        property_description = soup.find(
            'div', {'id': 'infodesctab'}).find('div', {'class': 'row-fluid'}).findAll('div', {'class': 'row-fluid'})
        for i in property_description:
            heading = i.find('h3')
            if heading.text.strip() == 'Current Year Inventory':
                finding_propertyinfo = i.findAll('table')
                for j in finding_propertyinfo:
                    caption = j.find('caption')
                    if caption:
                        if caption.text.strip() == 'Current Year Inventory':
                            actual_propertyinfo = j.find(
                                'tbody').findAll('tr')
                            for k in actual_propertyinfo:
                                datafound=k.find('td').text.strip()
                                if datafound == '':
                                    data_to_insert.append('0')
                                else:
                                    try:
                                        float(datafound)
                                        if datafound.find('.') != -1:
                                            data_to_insert.append(float(datafound))
                                        else:
                                            data_to_insert.append(int(datafound))
                                    except:
                                        data_to_insert.append(datafound)
    if notfound==False:
        save_path = "images"
        file_name = data[0]+'.jpg'
        file_path = os.path.join(save_path, file_name)
        with open(file_path, 'wb') as file:
            file.write(requests.get(img_link).content)
            print('Saved your image')
    write_data_to_excel('new_1.xlsx', 'Sheet1', data_to_insert, count)
    count+=1


driver.close()
    