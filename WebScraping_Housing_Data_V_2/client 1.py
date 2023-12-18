import openpyxl
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import undetected_chromedriver as uc
from bs4 import BeautifulSoup
import requests
import os
import csv

cho = Options()

Path = "C:\Projects\chromedriver.exe"

service = Service(Path)
cho.add_argument("--disable-dev-shm-usage")
cho.add_argument("disable-infobars")
cho.add_argument("--disable-extensions")
cho.add_argument("--no-sandbox")

driver = uc.Chrome(options=cho, service=service)


workbook = openpyxl.load_workbook('getting tax data for last 4 years on clients (1).xlsx')

sheet = workbook['Sheet1']
extracted_data=[]
print('Scanning Excel Sheet')
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        extracted_data.append(str(cell.value).replace(' ', ''))
print('Scanning Successfull')


def write_data_to_excel(file_path, sheet_name, data_to_insert, row):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    for column, value in enumerate(data_to_insert, start=1):
        sheet.cell(row=row, column=column, value=value)
    workbook.save(file_path)
    workbook.close()
    print('Saved to Excel Sheet')

count=2
for i, data in enumerate(extracted_data[count-2:]):
    print(i,data)
    driver.get(data)
    data_to_insert=[]
    found=0
    while(found==0):
        try:
            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            schooltax = soup.find('div', {'id': 'infotaxtab'}).findAll('table')
            taxablevalue=schooltax[0].find('tbody').findAll('tr')[1].findAll('td')[1].text.strip()
            secondtable=schooltax[1].find('tbody').findAll('tr')
            thirdtable=schooltax[2].find('tbody').findAll('tr')
            basic_star=''
            impact=''
            rate_top=''
            non_exempt=''
            st2024_40120=''
            general_tax_rate=''
            for i in secondtable:
                finding_schooltax=i.findAll('td')
                for j in finding_schooltax:
                    if j.text.strip()=='*40120':
                        st2024_40120=finding_schooltax[2].text.strip()
                    if j.text.strip()=='Basic Star':
                        basic_star=finding_schooltax[2].text.strip()
                        impact=finding_schooltax[3].text.replace(
                            '$', '').replace(',', '').strip()
            
            for i in thirdtable:
                finding_schooltax=i.findAll('td')
                for j in finding_schooltax:
                    if j.text.strip()=='Basic Star':
                        rate_top=finding_schooltax[1].text.strip()
                    if j.text.strip()=='Non-Exempt':
                        non_exempt=finding_schooltax[1].text.strip()
            
            data_to_insert.append(data)
            data_to_insert.append(taxablevalue)
            data_to_insert.append(st2024_40120)
            data_to_insert.append(basic_star)
            data_to_insert.append(impact)
            data_to_insert.append(rate_top)
            data_to_insert.append(non_exempt)
            data_to_insert.append('')
 
            driver.find_element(By.LINK_TEXT, 'General and School Taxes').click()
            driver.find_element(By.ID, 'selectyr').click()
            finding_general_tax = driver.find_element(
                By.ID, 'selectyr').find_elements(By.TAG_NAME, 'option')
            found=1
        except Exception as e:
            continue


    for i in finding_general_tax:
        basic_star=''
        impact=''
        rate_top=''
        non_exempt=''
        st2024_40120=''
        general_tax_rate=''
        if i.get_property('value') == '2023':
            i.click()
            found=0
            while(found==0):
                try:
                    page_source = driver.page_source
                    soup = BeautifulSoup(page_source, 'html.parser')
                    schooltax = soup.find('div', {'id': 'infotaxtab'}).findAll('table')
                    taxablevalue=schooltax[0].find('tbody').findAll('tr')[1].findAll('td')[1].text.strip()
                    secondtable=schooltax[1].find('tbody').findAll('tr')
                    thirdtable=schooltax[2].find('tbody').findAll('tr')
                    for i in secondtable:
                        finding_schooltax=i.findAll('td')
                        for j in finding_schooltax:
                            if j.text.strip()=='*40120':
                                st2024_40120=finding_schooltax[2].text.strip()
                            if j.text.strip()=='Basic Star':
                                basic_star=finding_schooltax[2].text.strip()
                                impact=finding_schooltax[3].text.replace(
                                    '$', '').replace(',', '').strip()
                    
                    for i in thirdtable:
                        finding_schooltax=i.findAll('td')
                        for j in finding_schooltax:
                            if j.text.strip()=='Basic Star':
                                rate_top=finding_schooltax[1].text.strip()
                            if j.text.strip()=='Non-Exempt':
                                non_exempt=finding_schooltax[1].text.strip()

                    data_to_insert.append(taxablevalue)
                    data_to_insert.append(st2024_40120)
                    data_to_insert.append(basic_star)
                    data_to_insert.append(impact)
                    data_to_insert.append(rate_top)
                    data_to_insert.append(non_exempt)
                    gentax = soup.find('div', {'id': 'gentax2023'}).find(
                        'table').find('tbody').findAll('tr')
                    general_tax=''
                    for j in gentax:
                        finding_total = j.findAll('td')
                        for k in finding_total:
                            if k.text.strip() == 'TOTAL':
                                general_tax_rate = finding_total[3].text.strip()
                                data_to_insert.append(general_tax_rate)
                    break
                
                except Exception as e:
                    continue
            break

    driver.find_element(By.ID, 'selectyr').click()
    finding_general_tax = driver.find_element(By.ID, 'selectyr').find_elements(By.TAG_NAME, 'option')
    for i in finding_general_tax:
        if i.get_property('value') == '2022':
            i.click()
            found=0
            while(found==0):
                try:
                    page_source = driver.page_source
                    soup = BeautifulSoup(page_source, 'html.parser')
                    schooltax = soup.find('div', {'id': 'infotaxtab'}).findAll('table')
                    taxablevalue=schooltax[0].find('tbody').findAll('tr')[1].findAll('td')[1].text.strip()
                    secondtable=schooltax[1].find('tbody').findAll('tr')
                    thirdtable=schooltax[2].find('tbody').findAll('tr')
                    for i in secondtable:
                        finding_schooltax=i.findAll('td')
                        for j in finding_schooltax:
                            if j.text.strip()=='*40120':
                                st2024_40120=finding_schooltax[2].text.strip()
                            if j.text.strip()=='Basic Star':
                                basic_star=finding_schooltax[2].text.strip()
                                impact=finding_schooltax[3].text.replace(
                                    '$', '').replace(',', '').strip()
                    
                    for i in thirdtable:
                        finding_schooltax=i.findAll('td')
                        for j in finding_schooltax:
                            if j.text.strip()=='Basic Star':
                                rate_top=finding_schooltax[1].text.strip()
                            if j.text.strip()=='Non-Exempt':
                                non_exempt=finding_schooltax[1].text.strip()
                    
                    data_to_insert.append(taxablevalue)
                    data_to_insert.append(st2024_40120)
                    data_to_insert.append(basic_star)
                    data_to_insert.append(impact)
                    data_to_insert.append(rate_top)
                    data_to_insert.append(non_exempt)
                    gentax = soup.find('div', {'id': 'gentax2022'}).find(
                        'table').find('tbody').findAll('tr')
                    general_tax=''
                    for j in gentax:
                        finding_total = j.findAll('td')
                        for k in finding_total:
                            if k.text.strip() == 'TOTAL':
                                general_tax_rate = finding_total[3].text.strip()
                                data_to_insert.append(general_tax_rate)    
                    break
                except Exception as e:
                    continue
            break

    driver.find_element(By.ID, 'selectyr').click()
    finding_general_tax = driver.find_element(By.ID, 'selectyr').find_elements(By.TAG_NAME, 'option')
    for i in finding_general_tax:
        if i.get_property('value') == '2021':
            i.click()
            found=0
            while(found==0):
                try:
                    page_source = driver.page_source
                    soup = BeautifulSoup(page_source, 'html.parser')
                    schooltax = soup.find('div', {'id': 'infotaxtab'}).findAll('table')
                    taxablevalue=schooltax[0].find('tbody').findAll('tr')[1].findAll('td')[1].text.strip()
                    secondtable=schooltax[1].find('tbody').findAll('tr')
                    thirdtable=schooltax[2].find('tbody').findAll('tr')
                    for i in secondtable:
                        finding_schooltax=i.findAll('td')
                        for j in finding_schooltax:
                            if j.text.strip()=='*40120':
                                st2024_40120=finding_schooltax[2].text.strip()
                            if j.text.strip()=='Basic Star':
                                basic_star=finding_schooltax[2].text.strip()
                                impact=finding_schooltax[3].text.replace(
                                    '$', '').replace(',', '').strip()
                    
                    for i in thirdtable:
                        finding_schooltax=i.findAll('td')
                        for j in finding_schooltax:
                            if j.text.strip()=='Basic Star':
                                rate_top=finding_schooltax[1].text.strip()
                            if j.text.strip()=='Non-Exempt':
                                non_exempt=finding_schooltax[1].text.strip()
                    
                    data_to_insert.append(taxablevalue)
                    data_to_insert.append(st2024_40120)
                    data_to_insert.append(basic_star)
                    data_to_insert.append(impact)
                    data_to_insert.append(rate_top)
                    data_to_insert.append(non_exempt)
                    gentax = soup.find('div', {'id': 'gentax2021'}).find(
                        'table').find('tbody').findAll('tr')
                    general_tax=''
                    for j in gentax:
                        finding_total = j.findAll('td')
                        for k in finding_total:
                            if k.text.strip() == 'TOTAL':
                                general_tax_rate = finding_total[3].text.strip()
                                data_to_insert.append(general_tax_rate)
                    
                    found=1
                    break
                except Exception as e:
                    continue
            break 

    print(data_to_insert)
    write_data_to_excel('client_data.xlsx', 'Sheet1', data_to_insert, count)
    count+=1


driver.close()
    